from flask import Flask, render_template, request, redirect, jsonify, session, url_for, flash, Response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from database import (
    get_db_connection, get_root_causes, get_filtered_defects, get_top_symptoms,
    get_top_items, get_pics, get_symptoms, get_ng_stations, get_related_stations,
    get_defected_items, get_categories, get_statuses,
    get_user_by_username, get_all_users, create_user_pending, update_user_role,
    delete_user, update_user_password, get_pending_users, approve_user, reject_user,
    get_daily_defects_summary, get_defect_by_id, update_defect,
    get_top_defects_chart_data, get_top_items_chart_data,
    get_category_chart_data, get_station_chart_data,
    get_models, add_model, delete_model, get_pending_users_count
)
import datetime
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'alsafy-xiaomi-defect-tracker-2025-secret'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ─── Recovery Code ────────────────────────────────────────────────
RECOVERY_CODE = "2332$$"

class User:
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

    def is_authenticated(self): return True
    def is_active(self): return True
    def is_anonymous(self): return False
    def get_id(self): return str(self.id)

@login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, username, role FROM users WHERE id = %s AND status = 'approved'", (user_id,))
    user_data = cur.fetchone()
    cur.close()
    conn.close()
    if user_data:
        return User(user_data['id'], user_data['username'], user_data['role'])
    return None

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if current_user.role not in roles:
                flash('Access denied. Insufficient permissions.', 'error')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_model_names():
    return [m['name'] for m in get_models()]

def get_notification_count():
    if current_user.is_authenticated and current_user.role == 'super_admin':
        return get_pending_users_count()
    return 0

# ─── Auth Routes ──────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user_data = cur.fetchone()
        cur.close()
        conn.close()
        if user_data:
            if user_data['status'] == 'pending':
                flash('Your account is pending approval by the administrator.', 'warning')
            elif user_data['status'] == 'rejected':
                flash('Your account registration was rejected. Contact admin.', 'error')
            elif check_password_hash(user_data['password_hash'], password):
                user = User(user_data['id'], user_data['username'], user_data['role'])
                login_user(user)
                return redirect(url_for('index'))
            else:
                flash('Invalid username or password', 'error')
        else:
            flash('Invalid username or password', 'error')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')
        if not username or not password:
            flash('Username and password are required.', 'error')
        elif password != confirm:
            flash('Passwords do not match.', 'error')
        elif len(password) < 6:
            flash('Password must be at least 6 characters.', 'error')
        else:
            if create_user_pending(username, generate_password_hash(password)):
                flash('Registration submitted! Please wait for admin approval.', 'success')
                return redirect(url_for('login'))
            else:
                flash('Username already exists. Please choose another.', 'error')
    return render_template('register.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    step = request.args.get('step', '1')
    if request.method == 'POST':
        code = request.form.get('recovery_code', '').strip()
        if code == RECOVERY_CODE:
            return redirect(url_for('reset_password'))
        else:
            flash('Invalid recovery code.', 'error')
    return render_template('forgot_password.html', step=step)

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        username  = request.form.get('username', '').strip()
        password  = request.form.get('new_password', '')
        confirm   = request.form.get('confirm_password', '')
        if password != confirm:
            flash('Passwords do not match.', 'error')
        elif len(password) < 6:
            flash('Password must be at least 6 characters.', 'error')
        else:
            conn = get_db_connection()
            cur  = conn.cursor()
            cur.execute("SELECT id FROM users WHERE username = %s", (username,))
            u = cur.fetchone()
            if u:
                update_user_password(u['id'], generate_password_hash(password))
                flash('Password reset successfully! Please login.', 'success')
                cur.close(); conn.close()
                return redirect(url_for('login'))
            else:
                flash('Username not found.', 'error')
            cur.close(); conn.close()
    return render_template('reset_password.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard'))

# ─── Dashboard ────────────────────────────────────────────────────

@app.route('/dashboard')
@login_required
def dashboard():
    model     = request.args.get('model')
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    sel_date  = request.args.get('date')
    q         = request.args.get('q','').strip()   # unified search

    if sel_date:
        date_from = sel_date
        date_to   = sel_date

    # Only load data when a filter is explicitly selected
    has_filter = bool(model or date_from or date_to or q)

    if has_filter:
        defects = get_filtered_defects(model, None, date_from, date_to,
                                       symptom_filter=q, item_filter=q, station_filter=q,
                                       unified_search=True)
        total_output = total_defects = 0
        for d in defects:
            if str(d.get('symptom', '')).upper() == 'TOTAL':
                total_output += (d.get('actual_out') or 0)
            else:
                total_defects += (d.get('defect_qty') or 0)
        defect_rate = (total_defects / total_output * 100) if total_output > 0 else 0
        ppm = defect_rate * 10000
        daily_summary    = get_daily_defects_summary(model, date_from, date_to)
        chart_defects    = get_top_defects_chart_data(model, date_from, date_to, 12)
        chart_items      = get_top_items_chart_data(model, date_from, date_to, 6)
        chart_categories = get_category_chart_data(model, date_from, date_to)
        chart_stations   = get_station_chart_data(model, date_from, date_to, 7)
    else:
        defects = []
        total_output = total_defects = 0
        defect_rate = ppm = 0
        daily_summary = chart_defects = chart_items = chart_categories = chart_stations = []

    stats = {
        "total_output":  total_output,
        "total_defects": total_defects,
        "defect_rate":   round(defect_rate, 2),
        "ppm":           round(ppm, 0)
    }

    notif_count = get_notification_count()
    worst1 = chart_defects[0] if len(chart_defects) > 0 else None
    worst2 = chart_defects[1] if len(chart_defects) > 1 else None

    from database import get_symptoms, get_ng_stations, get_defected_items
    return render_template('dashboard.html', view='dashboard', defects=defects, stats=stats,
        models=get_model_names(), selected_model=model,
        date_from=date_from, date_to=date_to, sel_date=sel_date,
        has_filter=has_filter,
        chart_defects=chart_defects, chart_items=chart_items,
        chart_categories=chart_categories, chart_stations=chart_stations,
        daily_summary=daily_summary, notif_count=notif_count,
        worst1=worst1, worst2=worst2,
        symptoms=get_symptoms(), ng_stations=get_ng_stations(), defected_items=get_defected_items(),
        user_role=current_user.role, username=current_user.username, datetime=datetime)

# ─── Defect Registration ─────────────────────────────────────────

@app.route('/defect-registration')
@login_required
def defect_registration():
    selected_model  = request.args.get('model')
    date_from       = request.args.get('date_from')
    date_to         = request.args.get('date_to')
    sel_date        = request.args.get('date')
    q               = request.args.get('q','').strip()   # unified search

    if sel_date:
        date_from = sel_date
        date_to   = sel_date

    # Only load records when a filter is explicitly selected
    has_filter = bool(selected_model or date_from or date_to or q)
    if has_filter:
        defects = get_filtered_defects(selected_model, None, date_from, date_to,
                                       symptom_filter=q, item_filter=q, station_filter=q,
                                       unified_search=True)
    else:
        defects = []

    notif_count = get_notification_count()
    # All authenticated users can add, edit and delete records.
    # Only super_admin/admin can manage users, models, bulk-delete, import/export.
    can_edit = True

    return render_template('registration.html', view='defect-registration',
        q=q,
        models=get_model_names(), models_full=get_models(), selected_model=selected_model,
        date_from=date_from, date_to=date_to, sel_date=sel_date,
        has_filter=has_filter,
        defects=defects,
        can_edit=can_edit,
        pics=get_pics(), symptoms=get_symptoms(), ng_stations=get_ng_stations(),
        related_stations=get_related_stations(), defected_items=get_defected_items(),
        categories=get_categories(), statuses=get_statuses(),
        root_causes=get_root_causes(), notif_count=notif_count,
        user_role=current_user.role, username=current_user.username, datetime=datetime)

@app.route('/add', methods=['POST'])
@login_required
def add_defect():
    data       = request.form
    symptom    = (data.get('symptom') or '').strip().upper()
    rec_date   = data.get('date') or str(datetime.today().date())
    model_val  = data.get('model')
    actual_out = int(data.get('actual_out', 0) or 0)
    actual_in  = int(data.get('actual_in',  0) or 0)
    target_in  = int(data.get('target_in',  0) or 0)
    target_out = int(data.get('target_out', 0) or 0)

    # For TOTAL rows: calculate defect% and ppm from day totals
    defect_pct = 0.0
    ppm_val    = 0.0
    defect_qty = 0

    if symptom == 'TOTAL':
        # Sum all non-TOTAL defect_qty for same model+date
        conn2 = get_db_connection(); cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COALESCE(SUM(defect_qty),0) FROM defects WHERE UPPER(symptom)!='TOTAL' AND date=%s AND model ILIKE %s",
            (rec_date, model_val)
        )
        row2 = cur2.fetchone(); cur2.close(); conn2.close()
        total_day_defects = int(list(row2.values())[0]) if row2 else 0
        if actual_out > 0:
            defect_pct = round((total_day_defects / actual_out) * 100, 6)
            ppm_val    = round(defect_pct * 10000, 2)
        defect_qty = 0
    else:
        defect_qty = int(data.get('defect_qty', 0) or 0)

    conn = get_db_connection()
    cur  = conn.cursor()
    cur.execute("""
    INSERT INTO defects (
        pic, date, model, shift, sn, symptom, ng_station, defect_qty, root_cause, related_station,
        defected_item, defected_item_sn, defected_item_qty, defect_pic, category, status, target_in,
        target_out, actual_in, actual_out, defect_percent, ppm, remarks
    ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data.get('pic') or None,
        rec_date,
        model_val,
        data.get('shift') or None,
        data.get('sn') or None,
        data.get('symptom') or None,
        data.get('ng_station') or None,
        defect_qty,
        data.get('root_cause') or None,
        data.get('related_station') or None,
        data.get('defected_item') or None,
        data.get('defected_item_sn') or None,
        int(data.get('defected_item_qty') or 1),
        data.get('defect_pic') or None,
        data.get('category') or None,
        data.get('status') or 'Pending',
        target_in, target_out, actual_in, actual_out,
        defect_pct, ppm_val,
        data.get('remarks', '')
    ))
    conn.commit()
    cur.close()
    conn.close()
    flash('Record saved successfully!', 'success')
    return redirect(url_for('defect_registration', model=model_val))

# ─── Edit / Delete Defect ────────────────────────────────────────

@app.route('/edit-defect/<int:defect_id>', methods=['GET', 'POST'])
@login_required
def edit_defect(defect_id):
    if request.method == 'POST':
        data       = request.form
        actual_out = int(data.get('actual_out', 0) or 0)
        conn = get_db_connection()
        cur  = conn.cursor()
        cur.execute("""
            UPDATE defects SET
                pic=%s,date=%s,model=%s,shift=%s,sn=%s,symptom=%s,ng_station=%s,defect_qty=%s,
                root_cause=%s,related_station=%s,defected_item=%s,defected_item_sn=%s,defected_item_qty=%s,
                defect_pic=%s,category=%s,status=%s,target_in=%s,target_out=%s,
                actual_in=%s,actual_out=%s,defect_percent=%s,ppm=%s,remarks=%s
            WHERE id=%s
        """, (
            data.get('pic'), data.get('date'), data.get('model'), data.get('shift'), data.get('sn'),
            data.get('symptom'), data.get('ng_station'), int(data.get('defect_qty') or 1),
            data.get('root_cause'), data.get('related_station'), data.get('defected_item'),
            data.get('defected_item_sn'), int(data.get('defected_item_qty') or 1),
            data.get('defect_pic'), data.get('category'), data.get('status') or 'Pending',
            int(data.get('target_in') or 0), int(data.get('target_out') or 0),
            int(data.get('actual_in') or 0), actual_out, 0.0, 0,
            data.get('remarks', ''), defect_id
        ))
        conn.commit()
        cur.close()
        conn.close()
        flash('Defect updated successfully!', 'success')
        return redirect(url_for('defect_registration', model=data.get('model')))

    defect = get_defect_by_id(defect_id)
    if not defect:
        return redirect(url_for('dashboard'))
    notif_count = get_notification_count()
    return render_template('edit_defect.html', view='edit-defect', defect=defect,
        models=get_model_names(), pics=get_pics(),
        symptoms=get_symptoms(), ng_stations=get_ng_stations(),
        related_stations=get_related_stations(),
        defected_items=get_defected_items(), categories=get_categories(),
        statuses=get_statuses(), root_causes=get_root_causes(),
        notif_count=notif_count,
        user_role=current_user.role, username=current_user.username, datetime=datetime)

@app.route('/delete-defect/<int:defect_id>', methods=['POST'])
@login_required
def delete_defect(defect_id):
    model = request.form.get('model')
    conn  = get_db_connection()
    cur   = conn.cursor()
    cur.execute("DELETE FROM defects WHERE id = %s", (defect_id,))
    conn.commit()
    cur.close()
    conn.close()
    flash('Defect deleted successfully!', 'success')
    return redirect(url_for('defect_registration', model=model))

# ─── User Management (super_admin only) ──────────────────────────

@app.route('/users')
@role_required('super_admin')
def users():
    notif_count = get_notification_count()
    return render_template('users.html', view='users',
        users=get_all_users(),
        pending_users=get_pending_users(),
        models=get_model_names(),
        notif_count=notif_count,
        user_role=current_user.role, username=current_user.username, datetime=datetime)

@app.route('/approve-user/<int:user_id>', methods=['POST'])
@role_required('super_admin')
def approve_user_route(user_id):
    role = request.form.get('role', 'user')
    approve_user(user_id, role)
    flash('User approved successfully!', 'success')
    return redirect(url_for('users'))

@app.route('/reject-user/<int:user_id>', methods=['POST'])
@role_required('super_admin')
def reject_user_route(user_id):
    reject_user(user_id)
    flash('User registration rejected.', 'info')
    return redirect(url_for('users'))

@app.route('/add-user', methods=['POST'])
@role_required('super_admin')
def add_user():
    username = request.form.get('username')
    password = request.form.get('password')
    role     = request.form.get('role', 'user')
    conn = get_db_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO users (username, password_hash, role, status) VALUES (%s,%s,%s,'approved')",
            (username, generate_password_hash(password), role)
        )
        conn.commit()
        flash(f'User {username} created successfully!', 'success')
    except Exception:
        conn.rollback()
        flash(f'Failed to create user. Username may already exist.', 'error')
    finally:
        cur.close(); conn.close()
    return redirect(url_for('users'))

@app.route('/update-user-role/<int:user_id>', methods=['POST'])
@role_required('super_admin')
def update_user_role_route(user_id):
    role = request.form.get('role')
    update_user_role(user_id, role)
    flash('User role updated successfully!', 'success')
    return redirect(url_for('users'))

@app.route('/delete-user/<int:user_id>', methods=['POST'])
@role_required('super_admin')
def delete_user_route(user_id):
    delete_user(user_id)
    flash('User deleted successfully!', 'success')
    return redirect(url_for('users'))

# ─── Model Management ─────────────────────────────────────────────

@app.route('/add-model', methods=['POST'])
@role_required('admin', 'super_admin')
def add_model_route():
    name = request.form.get('name', '').strip()
    if name:
        if add_model(name):
            flash(f'Model "{name}" added successfully!', 'success')
        else:
            flash(f'Model "{name}" already exists.', 'error')
    return redirect(url_for('defect_registration'))

@app.route('/delete-model/<int:model_id>', methods=['POST'])
@role_required('admin', 'super_admin')
def delete_model_route(model_id):
    delete_model(model_id)
    flash('Model deleted successfully!', 'success')
    return redirect(url_for('defect_registration'))

# ─── Export Excel ──────────────────────────────────────────────────

@app.route('/export-excel')
@role_required('admin', 'super_admin')
def export_excel():
    model     = request.args.get('model')
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    defects   = get_filtered_defects(model, None, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defects Report"

    # Header styling
    header_fill   = PatternFill("solid", fgColor="1a1a2e")
    header_font   = Font(color="FF6B35", bold=True, size=11)
    center_align  = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['ID','PIC','Date','Model','Shift','S/N','Symptom','NG Station',
               'Defect Qty','Root Cause','Related Station','Defected Item',
               'Defected Item S/N','Defected Item Qty','Defect PIC','Category',
               'Status','Target I/N','Target O/T','Actual I/N','Actual O/T',
               'Defect %','PPM','Remarks']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center_align
        cell.border = thin_border

    pending_fill = PatternFill("solid", fgColor="FFE0E0")
    total_fill   = PatternFill("solid", fgColor="E0FFE8")

    for row_idx, d in enumerate(defects, 2):
        values = [
            d.get('id'), d.get('pic'), str(d.get('date', '')), d.get('model'), d.get('shift'),
            d.get('sn'), d.get('symptom'), d.get('ng_station'), d.get('defect_qty'),
            d.get('root_cause'), d.get('related_station'), d.get('defected_item'),
            d.get('defected_item_sn'), d.get('defected_item_qty'), d.get('defect_pic'),
            d.get('category'), d.get('status'), d.get('target_in'), d.get('target_out'),
            d.get('actual_in'), d.get('actual_out'), d.get('defect_percent'), d.get('ppm'),
            d.get('remarks')
        ]
        is_pending = str(d.get('status', '')).lower() == 'pending'
        is_total   = str(d.get('symptom', '')).upper() == 'total'
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if is_total:
                cell.fill = total_fill
            elif is_pending:
                cell.fill = pending_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"defects_{model or 'all'}_{datetime.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={fname}'}
    )

# ─── Import Excel ──────────────────────────────────────────────────

@app.route('/import-excel', methods=['POST'])
@role_required('admin', 'super_admin')
def import_excel():
    if 'file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('defect_registration'))
    file = request.files['file']
    if not file or file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('defect_registration'))
    if not file.filename.lower().endswith(('.xlsx','.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('defect_registration'))
    try:
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(file.read()), data_only=True)
        ws = wb.active
        # Read headers from row 1 — strip whitespace
        raw = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(max_row=1))]
        # Map every known Excel header variation → db column name
        MAP = {
            'PIC':'pic','Date':'date','Model':'model','Shift':'shift',
            'N.G station':'ng_station','NG Station':'ng_station','NG station':'ng_station',
            'S/N':'sn','Symptom':'symptom','Defect Qty':'defect_qty','Defect QTY':'defect_qty',
            'Root Cause':'root_cause',
            'Related station':'related_station','Related Station':'related_station',
            'Defected Item':'defected_item',
            'Defected Item S/N':'defected_item_sn',
            'Defected Item QTY':'defected_item_qty','Defected Item Qty':'defected_item_qty',
            'Defect PIC':'defect_pic','Category':'category','Status':'status',
            'Target  I/N':'target_in','Target I/N':'target_in',
            'Target O/T':'target_out','Actual I/N':'actual_in','Actual O/T':'actual_out',
            'Defect %':'defect_percent','PPM':'ppm',
        }
        idx = {i: MAP[h] for i,h in enumerate(raw) if h in MAP}
        if not idx:
            flash('Could not find matching columns. Check your Excel file headers.', 'error')
            return redirect(url_for('defect_registration'))

        def si(v,d=0):
            try: return int(float(str(v))) if v is not None and str(v).strip() not in ('','None') else d
            except: return d
        def sf(v,d=0.0):
            try: return float(str(v)) if v is not None and str(v).strip() not in ('','None') else d
            except: return d
        def ss(v):
            if v is None: return None
            s=str(v).strip(); return s if s and s!='None' else None
        def sd(v):
            if v is None: return None
            return v.date() if hasattr(v,'date') else v

        conn = get_db_connection()
        cur  = conn.cursor()
        count = skipped = 0

        updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row[:16]): continue
            rd = {idx[i]: row[i] for i in idx if i < len(row)}
            if not rd.get('model') and not rd.get('symptom'): continue

            r_date  = sd(rd.get('date'))
            r_model = ss(rd.get('model'))
            r_sym   = ss(rd.get('symptom'))
            r_sn    = ss(rd.get('sn'))
            r_shift = ss(rd.get('shift'))

            # Auto-calculate defect% and ppm for TOTAL rows
            r_actual_out = si(rd.get('actual_out'))
            r_defect_pct = sf(rd.get('defect_percent'))
            r_ppm        = sf(rd.get('ppm'))
            if r_sym and r_sym.upper() == 'TOTAL' and r_actual_out > 0:
                # Count defects already inserted for this date/model
                cur.execute(
                    "SELECT COALESCE(SUM(defect_qty),0) as tot FROM defects WHERE UPPER(symptom)!='TOTAL' AND date=%s AND model ILIKE %s",
                    (r_date, r_model)
                )
                t = cur.fetchone()
                td = int(t['tot']) if t else 0
                r_defect_pct = round((td / r_actual_out) * 100, 6)
                r_ppm        = round(r_defect_pct * 10000, 2)

            try:
                # Only consider a row duplicate if EVERY identifying field matches exactly:
                # date + model + symptom + sn + shift + ng_station + defected_item + defect_qty
                # If ANY of these differ → it is a NEW record, always INSERT
                r_ng      = ss(rd.get('ng_station'))
                r_def_item= ss(rd.get('defected_item'))
                r_def_qty = si(rd.get('defect_qty'), 0)

                cur.execute("""
                    SELECT id FROM defects WHERE
                        date = %s
                        AND model ILIKE %s
                        AND UPPER(COALESCE(symptom,''))   = UPPER(COALESCE(%s,''))
                        AND COALESCE(sn,'')               = COALESCE(%s,'')
                        AND COALESCE(shift,'')            = COALESCE(%s,'')
                        AND COALESCE(ng_station,'')       = COALESCE(%s,'')
                        AND COALESCE(defected_item,'')    = COALESCE(%s,'')
                        AND COALESCE(defect_qty,0)        = %s
                    LIMIT 1
                """, (
                    r_date, r_model,
                    r_sym  or '', r_sn      or '',
                    r_shift or '', r_ng     or '',
                    r_def_item or '', r_def_qty
                ))
                existing = cur.fetchone()

                if existing:
                    # TRUE duplicate — update non-key fields only
                    cur.execute("""
                        UPDATE defects SET
                          pic=%s, root_cause=%s, related_station=%s,
                          defected_item_sn=%s, defected_item_qty=%s,
                          defect_pic=%s, category=%s, status=%s,
                          target_in=%s, target_out=%s, actual_in=%s, actual_out=%s,
                          defect_percent=%s, ppm=%s
                        WHERE id=%s
                    """, (
                        ss(rd.get('pic')), ss(rd.get('root_cause')),
                        ss(rd.get('related_station')), ss(rd.get('defected_item_sn')),
                        si(rd.get('defected_item_qty'), 1), ss(rd.get('defect_pic')),
                        ss(rd.get('category')), ss(rd.get('status')) or 'Pending',
                        si(rd.get('target_in')), si(rd.get('target_out')),
                        si(rd.get('actual_in')), r_actual_out,
                        r_defect_pct, r_ppm,
                        existing['id']
                    ))
                    updated += 1
                else:
                    # Different in at least one field → always INSERT as new row
                    cur.execute(
                        "INSERT INTO defects (pic,date,model,shift,sn,symptom,ng_station,defect_qty,"
                        "root_cause,related_station,defected_item,defected_item_sn,defected_item_qty,"
                        "defect_pic,category,status,target_in,target_out,actual_in,actual_out,"
                        "defect_percent,ppm,remarks) VALUES "
                        "(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (ss(rd.get('pic')), r_date, r_model, r_shift, r_sn, r_sym,
                         r_ng, r_def_qty,
                         ss(rd.get('root_cause')), ss(rd.get('related_station')),
                         r_def_item, ss(rd.get('defected_item_sn')),
                         si(rd.get('defected_item_qty'), 1), ss(rd.get('defect_pic')),
                         ss(rd.get('category')), ss(rd.get('status')) or 'Pending',
                         si(rd.get('target_in')), si(rd.get('target_out')),
                         si(rd.get('actual_in')), r_actual_out,
                         r_defect_pct, r_ppm, '')
                    )
                    count += 1
            except Exception as ex:
                conn.rollback(); skipped += 1
        conn.commit(); cur.close(); conn.close()
        parts = [f'Imported {count} new records']
        if updated: parts.append(f'{updated} updated')
        if skipped: parts.append(f'{skipped} skipped')
        flash(', '.join(parts) + '.', 'success')
    except Exception as e:
        flash(f'Import failed: {str(e)}', 'error')
    return redirect(url_for('defect_registration'))


# ─── Bulk Delete ──────────────────────────────────────────────────

@app.route('/delete-bulk', methods=['POST'])
@role_required('admin', 'super_admin')
def delete_bulk():
    mode       = request.form.get('mode')        # 'all' | 'date' | 'model'
    del_date   = request.form.get('del_date')
    del_model  = request.form.get('del_model')
    conn = get_db_connection()
    cur  = conn.cursor()
    if mode == 'all':
        cur.execute("DELETE FROM defects")
        conn.commit()
        flash('All records deleted successfully.', 'success')
    elif mode == 'date' and del_date:
        cur.execute("DELETE FROM defects WHERE date = %s", (del_date,))
        conn.commit()
        flash(f'All records for {del_date} deleted.', 'success')
    elif mode == 'model' and del_model:
        cur.execute("DELETE FROM defects WHERE model ILIKE %s", (del_model,))
        conn.commit()
        flash(f'All records for model {del_model} deleted.', 'success')
    else:
        flash('Invalid delete request.', 'error')
    cur.close(); conn.close()
    return redirect(url_for('defect_registration'))

@app.route('/api/dashboard-stats')
@login_required
def api_dashboard_stats():
    model     = request.args.get('model')
    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')

    daily       = get_daily_defects_summary(model, date_from, date_to)
    top_defects = get_top_defects_chart_data(model, date_from, date_to, 3)
    top_items   = get_top_items_chart_data(model, date_from, date_to, 3)

    daily_data = []
    for d in daily:
        output  = int(d['total_output'] or 0)
        defects = int(d['total_defects'] or 0)
        rate    = (defects / output * 100) if output > 0 else 0
        daily_data.append({
            'date': str(d['date']),
            'total_output': output,
            'total_defects': defects,
            'defect_rate': round(rate, 2)
        })

    return jsonify({
        'daily': daily_data,
        'top_defects': [{'symptom': t['symptom'], 'qty': int(t['total_qty'])} for t in top_defects],
        'top_items':   [{'item': t['defected_item'], 'qty': int(t['total_qty'])} for t in top_items]
    })

@app.route('/api/notifications')
@role_required('super_admin')
def api_notifications():
    return jsonify({'count': get_pending_users_count()})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5002, debug=False)
